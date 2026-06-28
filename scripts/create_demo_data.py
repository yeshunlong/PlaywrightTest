import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

headers = [
    "1级目录", "用例名称", "需求id", "备注", "大前提",
    "步骤名称", "前置条件", "操作描述", "参数", "预期结果"
]

data = [
    [
        "布局与显示", "TC_组件窗口_基础布局", "", "",
        "客户端已登录，已打开组件入口", "窗口整体布局", "",
        "1、打开【通用工具】-【示例板块】组件\n2、观察窗口整体区域", "",
        "窗口上方显示标题栏；中间显示内容区；右上角显示菜单按钮"
    ],
    [
        "布局与显示", "TC_组件窗口_默认尺寸", "", "",
        "客户端已登录，已打开组件入口", "默认尺寸", "",
        "1、打开【通用工具】-【示例板块】组件\n2、观察组件初始窗口大小", "",
        "窗口以默认尺寸打开，内容区域完整显示，无明显遮挡"
    ],
    [
        "布局与显示", "TC_导航栏_标签显示", "", "",
        "已打开示例板块组件", "标签导航栏", "",
        "1、查看组件顶部标签导航栏\n2、切换任一标签后再切回默认标签", "",
        "导航栏展示多个业务标签；当前选中标签有高亮状态；切换后内容区刷新"
    ],
    [
        "布局与显示", "TC_列表区域_基础显示", "", "",
        "已打开示例板块组件", "列表区域", "",
        "1、查看左侧列表区域\n2、点击左侧任一分类\n3、查看右侧明细区域", "",
        "左侧显示分类列表，右侧显示对应明细；切换分类后明细区域随之刷新"
    ],
    [
        "布局与显示", "TC_窗口缩放_横向拉伸", "", "",
        "已打开示例板块组件", "横向拉伸", "",
        "1、拖动窗口右边缘向右拉伸\n2、观察内容区域变化", "",
        "窗口宽度变大后，内容区域随窗口自适应扩展，主要控件不重叠"
    ],
    [
        "布局与显示", "TC_组件入口_打开板块", "", "",
        "客户端已登录", "打开组件", "",
        "1、打开示例板块", "",
        "弹出示例板块组件窗口"
    ],
    [
        "布局与显示", "TC_导航栏_默认选中项", "", "",
        "已打开示例板块组件", "默认选中", "",
        "1、查看顶部导航栏默认选中的标签", "",
        "默认选中【地域指数】标签"
    ],
    [
        "布局与显示", "TC_右上角菜单_分组设置", "", "",
        "已打开示例板块组件", "菜单项检查", "",
        "1、点击菜单\n2、查看是否有分组设置", "",
        "菜单中显示【设置组件分组】选项"
    ],
    [
        "布局与显示", "TC_窗口最小宽度_校验", "", "",
        "已打开示例板块组件", "最小宽度", "",
        "1、横向压缩组件窗口直到不能继续压缩\n2、观察窗口状态", "",
        "窗口达到最小宽度"
    ],
    [
        "布局与显示", "TC_列表刷新_切换分类", "", "",
        "已打开示例板块组件", "列表刷新", "",
        "1、点击左侧任意分类\n2、观察右侧明细列表", "",
        "右侧明细列表颜色、数量、排序均与行情中心保持一致"
    ],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "布局与显示"

header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
cell_alignment = Alignment(vertical="top", wrap_text=True)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        cell.font = Font(name="微软雅黑", size=10)

column_widths = {1: 12, 2: 26, 3: 10, 4: 10, 5: 30, 6: 14, 7: 10, 8: 45, 9: 10, 10: 50}
for col, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

for row_idx in range(2, 12):
    ws.row_dimensions[row_idx].height = 60

output_path = r"C:\Data\Document\社招\CV\杭州\同花顺\同花顺面试题\PlaywrightTest\data\布局与显示.xlsx"
wb.save(output_path)
print(f"File saved to {output_path}")
