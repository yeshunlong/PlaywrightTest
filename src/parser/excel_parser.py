"""
Excel 测试用例解析器

负责读取原始 Excel 文件，解析测试用例的字段映射，
将每行数据转换为结构化的 TestCase 模型。

支持的 Excel 格式：
- 布局与显示.xls 的标准列：1级目录、用例名称、需求id、备注、大前提、
  步骤名称、前置条件、操作描述、参数、预期结果
"""

from pathlib import Path
from typing import Optional

import openpyxl
from loguru import logger

from src import TestCase


class ExcelParser:
    """Excel 用例解析器"""

    # 列名到字段名的映射
    COLUMN_MAPPING = {
        "1级目录": "category",
        "用例名称": "name",
        "需求id": "requirement_id",
        "备注": "remarks",
        "大前提": "macro_precondition",
        "步骤名称": "step_type",
        "前置条件": "inner_precondition",
        "操作描述": "operation_desc",
        "参数": "params",
        "预期结果": "expected_result",
    }

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.file_path}")

    def parse(self) -> list[TestCase]:
        """解析 Excel 文件，返回测试用例列表"""
        logger.info(f"开始解析 Excel: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb.active
        logger.info(f"工作表: {ws.title}, 行数: {ws.max_row}, 列数: {ws.max_column}")

        # 读取表头
        headers = self._read_headers(ws)
        logger.info(f"检测到列: {headers}")

        # 建立列索引映射
        col_map = self._build_column_map(headers)
        logger.info(f"列映射: {col_map}")

        # 解析数据行
        cases = []
        for row_idx in range(2, ws.max_row + 1):
            raw_data = {}
            for col_idx, header in enumerate(headers, 1):
                raw_data[header] = ws.cell(row=row_idx, column=col_idx).value or ""

            # 跳过完全空行
            if all(not v for v in raw_data.values()):
                continue

            case = self._row_to_case(raw_data, row_idx)
            if case:
                cases.append(case)
                logger.debug(f"  解析用例 [{row_idx}]: {case.name}")

        logger.info(f"解析完成，共 {len(cases)} 条用例")
        wb.close()
        return cases

    def _read_headers(self, ws) -> list[str]:
        """读取第一行作为表头"""
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            headers.append(str(val).strip() if val else f"col_{col_idx}")
        return headers

    def _build_column_map(self, headers: list[str]) -> dict[str, int]:
        """建立字段名到列索引的映射"""
        col_map = {}
        for idx, header in enumerate(headers):
            # 尝试匹配标准字段名
            for key, field in self.COLUMN_MAPPING.items():
                if key in header:
                    col_map[field] = idx
                    break
        return col_map

    def _row_to_case(self, raw_data: dict, row_idx: int) -> Optional[TestCase]:
        """将一行原始数据转换为 TestCase"""
        data = {}
        for header, value in raw_data.items():
            # 匹配字段
            for key, field in self.COLUMN_MAPPING.items():
                if key in str(header):
                    data[field] = str(value).strip() if value else ""
                    break

        # 生成 case_id（如果没有则从名称生成）
        name = data.get("name", f"TC_ROW_{row_idx}")
        case_id = data.get("case_id", name)

        return TestCase(
            case_id=case_id,
            name=name,
            category=data.get("category", ""),
            requirement_id=data.get("requirement_id", ""),
            remarks=data.get("remarks", ""),
            macro_precondition=data.get("macro_precondition", ""),
            inner_precondition=data.get("inner_precondition", ""),
            step_type=data.get("step_type", ""),
            operation_desc=data.get("operation_desc", ""),
            params=data.get("params", ""),
            expected_result=data.get("expected_result", ""),
            raw_data=raw_data,
        )

    def parse_to_dict_list(self) -> list[dict]:
        """解析为字典列表（用于 AI prompt 的原始传递）"""
        cases = self.parse()
        return [
            {
                "序号": i + 1,
                "用例编号": c.case_id,
                "用例名称": c.name,
                "一级目录": c.category,
                "大前提": c.macro_precondition,
                "前置条件": c.inner_precondition,
                "操作步骤": c.operation_desc,
                "预期结果": c.expected_result,
                "步骤类型": c.step_type,
                "参数": c.params,
            }
            for i, c in enumerate(cases)
        ]
