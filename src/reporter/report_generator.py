"""
报告生成模块

生成以下输出文件：
1. cleaned_cases.xlsx  - 清洗后的测试用例（在原格式基础上增加清洗结果列）
2. cleaning_report.md  - 清洗报告（Markdown）
3. cleaning_report.html - 清洗报告（HTML）

报告包含：
- 执行统计
- 标签分布
- 逐用例详细分析
- 修复建议
- 可视化图表数据
"""

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from src import CleaningReport, CaseExecutionResult


class ReportGenerator:
    """报告生成器"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    PASSED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAILED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    NORMAL_FONT = Font(name="微软雅黑", size=10)
    TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_all(self, report: CleaningReport) -> dict[str, Path]:
        """生成所有报告文件"""
        results: dict[str, Path] = {}

        logger.info("生成清洗报告...")

        # 1. 生成 Excel
        excel_path = self.generate_excel(report)
        results["excel"] = excel_path
        logger.info(f"  Excel 报告: {excel_path}")

        # 2. 生成 Markdown
        md_path = self.generate_markdown(report)
        results["markdown"] = md_path
        logger.info(f"  Markdown 报告: {md_path}")

        # 3. 生成 HTML
        html_path = self.generate_html(report)
        results["html"] = html_path
        logger.info(f"  HTML 报告: {html_path}")

        return results

    def generate_excel(self, report: CleaningReport) -> Path:
        """生成清洗后的 Excel 文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "清洗结果"

        # 定义列
        headers = [
            "序号", "用例编号", "用例名称", "1级目录", "步骤名称",
            "操作描述", "预期结果", "是否通过", "未通过原因",
            "清洗标签", "AI修复建议"
        ]

        # 写表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.THIN_BORDER

        # 写数据
        for row_idx, result in enumerate(report.results, 2):
            case = result.case
            data = [
                row_idx - 1,           # 序号
                case.case_id,          # 用例编号
                case.name,             # 用例名称
                case.category,         # 1级目录
                case.step_type,        # 步骤名称
                case.operation_desc,   # 操作描述
                case.expected_result,  # 预期结果
                "是" if result.passed else "否",  # 是否通过
                result.reason if not result.passed else "",  # 未通过原因
                ", ".join(result.tags),  # 清洗标签
                result.ai_suggestion,    # AI修复建议
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.NORMAL_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = self.THIN_BORDER

                # 根据是否通过着色
                if result.passed:
                    cell.fill = self.PASSED_FILL
                else:
                    cell.fill = self.FAILED_FILL

        # 设置列宽
        col_widths = [6, 25, 22, 12, 15, 45, 40, 10, 40, 25, 50]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 第二个 sheet: 统计分析
        ws2 = wb.create_sheet("统计分析")

        stats_data = [
            ["统计项", "数值", "占比"],
            ["总用例数", report.total_cases, "100%"],
            ["通过数", report.passed_count, f"{report.passed_count/report.total_cases*100:.1f}%" if report.total_cases else "N/A"],
            ["未通过数", report.failed_count, f"{report.failed_count/report.total_cases*100:.1f}%" if report.total_cases else "N/A"],
            ["其中: 不可执行", report.not_executable_count, f"{report.not_executable_count/report.total_cases*100:.1f}%" if report.total_cases else "N/A"],
        ]

        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                else:
                    cell.font = self.NORMAL_FONT
                cell.border = self.THIN_BORDER

        # 标签分布
        row_offset = len(stats_data) + 2
        ws2.cell(row=row_offset, column=1, value="标签分布").font = Font(
            name="微软雅黑", size=12, bold=True
        )

        for i, (tag, count) in enumerate(report.tag_statistics.items()):
            r = row_offset + 1 + i
            ws2.cell(row=r, column=1, value=tag).font = self.NORMAL_FONT
            ws2.cell(row=r, column=2, value=count).font = self.NORMAL_FONT

        path = self.output_dir / "cleaned_cases.xlsx"
        wb.save(path)
        return path

    def generate_markdown(self, report: CleaningReport) -> Path:
        """生成 Markdown 报告"""
        lines = [
            f"# {report.title}",
            "",
            f"**生成时间**: {report.generated_at}",
            "",
            "---",
            "",
            "## 1. 总体统计",
            "",
            f"| 指标 | 数值 | 占比 |",
            f"|------|------|------|",
            f"| 总用例数 | {report.total_cases} | 100% |",
        ]

        if report.total_cases > 0:
            rate = report.passed_count / report.total_cases * 100
            lines.append(f"| 通过数 | {report.passed_count} | {rate:.1f}% |")
            lines.append(f"| 未通过数 | {report.failed_count} | {100-rate:.1f}% |")
        else:
            lines.append("| 通过数 | 0 | N/A |")
            lines.append("| 未通过数 | 0 | N/A |")

        lines.extend([
            f"| 其中: 不可执行 | {report.not_executable_count} | - |",
            "",
            "## 2. 标签分布",
            "",
            "| 标签 | 数量 |",
            "|------|------|",
        ])

        for tag, count in report.tag_statistics.items():
            lines.append(f"| {tag} | {count} |")

        lines.extend([
            "",
            "## 3. AI 摘要分析",
            "",
            report.summary or "_（未生成摘要）_",
            "",
            "## 4. 逐用例详情",
            "",
        ])

        for i, result in enumerate(report.results, 1):
            case = result.case
            status = "✅ 通过" if result.passed else "❌ 未通过"

            lines.extend([
                f"### {i}. {case.name} {status}",
                "",
                f"- **用例编号**: {case.case_id}",
                f"- **操作描述**: {case.operation_desc}",
                f"- **预期结果**: {case.expected_result}",
                f"- **清洗标签**: {', '.join(result.tags)}",
                f"- **判断原因**: {result.reason}",
            ])

            if result.ai_suggestion:
                lines.extend([
                    f"- **修复建议**:",
                ])
                for sug_line in result.ai_suggestion.split("\n"):
                    lines.append(f"  {sug_line}")

            lines.append("")

        lines.extend([
            "## 5. 修复建议汇总",
            "",
        ])
        if report.suggestions:
            for sug in report.suggestions:
                lines.append(f"- {sug}")
        else:
            lines.append("_暂无修复建议_")

        lines.extend([
            "",
            "---",
            "",
            "*此报告由 AI 测试用例清洗工程自动生成*",
        ])

        path = self.output_dir / "cleaning_report.md"
        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    def generate_html(self, report: CleaningReport) -> Path:
        """生成 HTML 报告"""
        passed_rate = (
            f"{report.passed_count / report.total_cases * 100:.1f}%"
            if report.total_cases > 0
            else "N/A"
        )

        # 构建用例行
        case_rows = ""
        for result in report.results:
            case = result.case
            status_class = "passed" if result.passed else "failed"
            status_text = "通过" if result.passed else "未通过"
            badge_color = "#22c55e" if result.passed else "#ef4444"

            tag_badges = " ".join(
                f'<span style="background:#e2e8f0;padding:2px 8px;border-radius:4px;font-size:12px;margin:2px">{tag}</span>'
                for tag in result.tags
            )

            suggestion_html = ""
            if result.ai_suggestion:
                suggestion_html = (
                    '<div style="margin-top:8px;padding:8px;background:#fef3c7;border-radius:4px;font-size:13px">'
                    f'<strong>修复建议:</strong><br>{result.ai_suggestion.replace(chr(10), "<br>")}'
                    "</div>"
                )

            case_rows += f"""
            <tr class="{status_class}">
                <td>{case.case_id}</td>
                <td>{case.name}</td>
                <td>{case.operation_desc[:80]}...</td>
                <td>{case.expected_result[:80]}...</td>
                <td><span class="badge" style="background:{badge_color}">{status_text}</span></td>
                <td style="font-size:12px">{result.reason}</td>
                <td>{tag_badges}{suggestion_html}</td>
            </tr>
            """

        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report.title}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Microsoft YaHei', sans-serif; background: #f0f2f5; color: #333; }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #1f4e79, #2d6da5); color: white; padding: 30px; border-radius: 8px; margin-bottom: 20px; }}
        .header h1 {{ font-size: 24px; }}
        .header p {{ opacity: 0.8; margin-top: 8px; }}
        .stats {{ display: flex; gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }}
        .stat-card {{ background: white; padding: 20px; border-radius: 8px; flex: 1; min-width: 180px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center; }}
        .stat-card .value {{ font-size: 32px; font-weight: bold; }}
        .stat-card .label {{ color: #666; font-size: 14px; margin-top: 4px; }}
        .stat-card.total .value {{ color: #1f4e79; }}
        .stat-card.passed .value {{ color: #22c55e; }}
        .stat-card.failed .value {{ color: #ef4444; }}
        .section {{ background: white; border-radius: 8px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .section h2 {{ font-size: 18px; margin-bottom: 16px; color: #1f4e79; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
        th {{ background: #1f4e79; color: white; padding: 10px 12px; text-align: left; }}
        td {{ padding: 10px 12px; border-bottom: 1px solid #e2e8f0; }}
        tr.passed {{ background: #f0fdf4; }}
        tr.failed {{ background: #fef2f2; }}
        .badge {{ color: white; padding: 2px 8px; border-radius: 4px; font-size: 12px; }}
        .summary {{ line-height: 1.8; white-space: pre-wrap; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{report.title}</h1>
            <p>生成时间: {report.generated_at}</p>
        </div>

        <div class="stats">
            <div class="stat-card total">
                <div class="value">{report.total_cases}</div>
                <div class="label">总用例数</div>
            </div>
            <div class="stat-card passed">
                <div class="value">{report.passed_count}</div>
                <div class="label">通过</div>
            </div>
            <div class="stat-card failed">
                <div class="value">{report.failed_count}</div>
                <div class="label">未通过</div>
            </div>
            <div class="stat-card">
                <div class="value">{passed_rate}</div>
                <div class="label">通过率</div>
            </div>
        </div>

        <div class="section">
            <h2>标签分布</h2>
            <table>
                <tr><th>标签</th><th>数量</th></tr>
                {"".join(f"<tr><td>{tag}</td><td>{count}</td></tr>" for tag, count in report.tag_statistics.items())}
            </table>
        </div>

        <div class="section">
            <h2>AI 摘要分析</h2>
            <div class="summary">{report.summary or "<em>（未生成摘要）</em>"}</div>
        </div>

        <div class="section">
            <h2>逐用例详情</h2>
            <table>
                <tr>
                    <th style="width:140px">用例编号</th>
                    <th style="width:160px">用例名称</th>
                    <th>操作描述</th>
                    <th>预期结果</th>
                    <th style="width:70px">状态</th>
                    <th style="width:200px">原因</th>
                    <th style="width:250px">标签 / 建议</th>
                </tr>
                {case_rows}
            </table>
        </div>

        <div style="text-align:center;color:#999;font-size:12px;padding:20px">
            此报告由 AI 测试用例清洗工程自动生成
        </div>
    </div>
</body>
</html>
"""
        path = self.output_dir / "cleaning_report.html"
        path.write_text(html, encoding="utf-8")
        return path
